#!/usr/bin/env python3
"""Create formatted Excel file from DB_enriched.csv"""
import pandas as pd
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read CSV
df = pd.read_csv('DB_enriched.csv', encoding='utf-8')
print(f"Read {len(df)} rows, {len(df.columns)} columns")

# Replace 'not_mentioned' with more readable text
df = df.fillna('not_mentioned')

# Save to Excel
output = 'DB_enriched.xlsx'
df.to_excel(output, index=False, sheet_name='NanoPlacentaDB')

# Format the Excel file
wb = load_workbook(output)
ws = wb['NanoPlacentaDB']

# Header styling
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx in range(1, len(df.columns) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Color-code material rows
material_colors = {
    'Au': 'FFF2CC',        # Gold
    'Ag': 'D9E2F3',       # Silver-blue
    'liposome': 'E2EFDA',  # Green
    'LNP': 'D5F5E3',      # Light green
    'lipid_NP': 'D5F5E3',
    'TiO2': 'FCE4EC',     # Pink
    'SiO2': 'F3E5F5',     # Purple
    'polystyrene': 'FFF3E0', # Orange
    'PLGA': 'E8F5E9',     # Mint
}

mat_col = list(df.columns).index('core_material') + 1

for row_idx in range(2, len(df) + 2):
    mat_val = ws.cell(row=row_idx, column=mat_col).value
    color = material_colors.get(str(mat_val), None)
    if color:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

# Auto-width columns (with max)
for col_idx in range(1, len(df.columns) + 1):
    max_len = len(str(ws.cell(row=1, column=col_idx).value))
    for row_idx in range(2, min(len(df) + 2, 50)):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            max_len = max(max_len, min(len(str(val)), 40))
    ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

# Freeze top row and first 2 columns
ws.freeze_panes = 'C2'

# Add autofilter
ws.auto_filter.ref = ws.dimensions

# Add summary sheet
ws2 = wb.create_sheet('Summary')
ws2['A1'] = 'NanoPlacentaDB - Enriched Summary'
ws2['A1'].font = Font(bold=True, size=14)

summary_data = [
    ('', ''),
    ('Total rows', len(df)),
    ('Unique studies', df['study_id'].nunique()),
    ('Unique PMIDs', df['pmid'].nunique()),
    ('Year range', f"{df['year'].min()} - {df['year'].max()}"),
    ('', ''),
    ('--- Material Breakdown ---', ''),
    ('Gold (Au)', len(df[df['core_material'] == 'Au'])),
    ('Liposome', len(df[df['core_material'] == 'liposome'])),
    ('LNP', len(df[df['core_material'] == 'LNP'])),
    ('Polystyrene', len(df[df['core_material'] == 'polystyrene'])),
    ('Silver (Ag)', len(df[df['core_material'] == 'Ag'])),
    ('TiO2', len(df[df['core_material'] == 'TiO2'])),
    ('SiO2', len(df[df['core_material'] == 'SiO2'])),
    ('PLGA', len(df[df['core_material'] == 'PLGA'])),
    ('Other/Unspecified', len(df[~df['core_material'].isin(['Au','liposome','LNP','polystyrene','Ag','TiO2','SiO2','PLGA'])])),
    ('', ''),
    ('--- Focus Areas (Shiri) ---', ''),
    ('Gold NP studies', len(df[df['core_material'] == 'Au'])),
    ('Liposome/LNP studies', len(df[df['core_material'].isin(['liposome', 'LNP', 'lipid_NP'])])),
    ('Combined Au + Lipo/LNP', len(df[df['core_material'].isin(['Au', 'liposome', 'LNP', 'lipid_NP'])])),
    ('', ''),
    ('--- Confidence ---', ''),
    ('High', len(df[df['confidence'] == 'high'])),
    ('Medium', len(df[df['confidence'] == 'medium'])),
    ('Low', len(df[df['confidence'] == 'low'])),
]

for i, (label, value) in enumerate(summary_data, start=2):
    ws2[f'A{i}'] = label
    ws2[f'B{i}'] = value
    if '---' in str(label):
        ws2[f'A{i}'].font = Font(bold=True)

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 15

wb.save(output)
print(f"\nSaved formatted Excel to {output}")
print(f"  Sheet 1: NanoPlacentaDB ({len(df)} rows)")
print(f"  Sheet 2: Summary statistics")
